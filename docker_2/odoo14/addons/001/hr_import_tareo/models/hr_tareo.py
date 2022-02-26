# -*- coding: utf-8 -*-
##############################################################################

from odoo import api, fields, models
from odoo.exceptions import Warning
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Protection, Font
import base64
from xlrd import open_workbook, xldate_as_tuple
from datetime import time
import os

class hr_tareo(models.Model):
    _inherit = 'hr.tareo'

    import_file = fields.Binary(string='Archivo a importar')
    tareo_load = fields.Boolean(string='Novedades importado')

    def import_tareo(self):
        self.ensure_one()
        if self.import_file and self.country_id:
            wb = open_workbook(file_contents = base64.decodestring(self.import_file))
            sheet = wb.sheet_by_index(0)
            for row in range(1,sheet.nrows):
                if sheet.cell(row,0).value:
                    dni = str(int(sheet.cell(row,0).value))
                    tareo_line_id = self.env['hr.tareo.detail'].search([('doc_number','=',dni),('tareo_id','=',self.id)])
                    print("DNI", dni, type(dni))
                    if tareo_line_id and self.country_id.code == 'PE':
                        tareo_line_id.update({
                            'ot25':float(sheet.cell(row,1).value * 24 * 3600)/ 3600,
                            'ot35':float(sheet.cell(row,2).value * 24 * 3600)/ 3600,
                            'ot100':sheet.cell(row,3).value,
                            'number_leave':int(sheet.cell(row,4).value or 0),
                            'hours_of_delay':float(sheet.cell(row,5).value * 24 * 3600)/60,
                            'holidays':int(sheet.cell(row,6).value or 0),
                            'holiday_sale':int(sheet.cell(row,7).value or 0),
                            'medical_breaks':int(sheet.cell(row,8).value or 0),
                            'maternity_allowance':int(sheet.cell(row,9).value or 0),
                            'sickness_allowance':int(sheet.cell(row,10).value or 0),
                            'license_with_enjoy':int(sheet.cell(row,11).value or 0),
                            'leave_without_enjoyment':int(sheet.cell(row,12).value or 0)
                            })
                    elif tareo_line_id and self.country_id.code == 'CO':
                        tareo_line_id.update({
                            'hed_co':int(sheet.cell(row,1).value or 0),
                            'hen_co':int(sheet.cell(row,2).value or 0),
                            'hefd_co':int(sheet.cell(row,3).value or 0),
                            'hefn_co':int(sheet.cell(row,4).value or 0),
                            'refe_co':int(sheet.cell(row,5).value or 0),
                            'reno_co':int(sheet.cell(row,6).value or 0),
                            'renf_co':int(sheet.cell(row,7).value or 0),
                            'fesc_co':int(sheet.cell(row,8).value or 0),
                            'ige_co':int(sheet.cell(row,9).value or 0),
                            'irl_co':int(sheet.cell(row,10).value or 0),
                            'lma_co':int(sheet.cell(row,11).value or 0),
                            'lpa_co':int(sheet.cell(row,12).value or 0),
                            'vco_co':int(sheet.cell(row,13).value or 0),
                            'vdi_co':int(sheet.cell(row,14).value or 0),
                            'vre_co':int(sheet.cell(row,15).value or 0),
                            'lnr_co':int(sheet.cell(row,16).value or 0),
                            'sln_co':int(sheet.cell(row,17).value or 0),
                            'lr_co':int(sheet.cell(row,18).value or 0),
                            'lt_co':int(sheet.cell(row,19).value or 0),
                            })
                    else:
                        raise Warning('No se encuentra una linea de Novedades para el documento %s') % str(dni)
            self.write({'tareo_load':True})

    def download_template(self):

        name = 'plantilla_novedades'

        return {
            'name': "Plantilla Novedades",
            'type': 'ir.actions.act_url',
            'url': self.env['ir.config_parameter'].get_param('web.base.url')+'/hr_import_tareo/static/src/archivos/'+name+'.xlsx',
            'target': 'new',
        }