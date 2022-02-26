# -*- coding: utf-8 -*-
from odoo import models, fields, api
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Protection, Font
import os
from odoo.exceptions import ValidationError
import xml.etree.ElementTree as ET


def as_text(value): 
    if value is None: 
     return "" 
    if type(value)=='float':
        return str(truncate(value,2))
    return str(value)



class HRWorkEntryType(models.Model):
    _inherit = 'hr.work.entry.type'

    overtime = fields.Boolean(string='Sobretiempo')
    ordinary_day = fields.Boolean(string='Jornada Ordinaria')

class HRSalaryRule(models.Model):
    _inherit = 'hr.salary.rule'
    type_concept = fields.Selection(

    string="Tipo de concepto",
    selection=[
          ('entry', 'Ingresos'),
          ('discount', 'Descuentos'),
          ('worker_contribution', 'Aportes del Trabajador'),
          ('employer_contribution', 'Aportes del Empleador'),
    ], default="entry"
    )  
    
class HrPayslipRun(models.Model):
    _inherit = 'hr.payslip.run'

    def export_report(self):
        RUTA_BASE = os.path.dirname(os.path.abspath(__file__))
        wb = Workbook()
        ws = wb.active
        bd = Side(style='thin', color="000000")
        normal_style = NamedStyle(name="normal_style")
        normal_style.font = Font(name='Arial Narrow', bold=False, size=10)
        normal_style.alignment = Alignment(horizontal='right')
        wb.add_named_style(normal_style)
        normal_style_border = NamedStyle(name="normal_style_border")
        normal_style_border.font = Font(name='Arial Narrow', bold=False, size=10)
        normal_style_border.alignment = Alignment(horizontal='right')
        normal_style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        wb.add_named_style(normal_style_border)
        black_style = NamedStyle(name="black_style")
        black_style.font = Font(name='Arial Narrow', bold=True, size=10)
        black_style.alignment = Alignment(horizontal='right')
        wb.add_named_style(black_style)
        black_style_border = NamedStyle(name="black_style_border")
        black_style_border.font = Font(name='Arial Narrow', bold=True, size=10)
        black_style_border.alignment = Alignment(horizontal='right')
        black_style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        wb.add_named_style(black_style_border)

        #Cabeceras
        row = 1
        rows = str(row)

        ws['A'+rows] = 'Centro de Costo'
        ws['A'+rows].style = black_style_border
        ws['A'+rows].alignment = Alignment(horizontal='center')
        ws['B'+rows] = 'Cod Trabajador'
        ws['B'+rows].style = black_style_border
        ws['B'+rows].alignment = Alignment(horizontal='center')
        ws['C'+rows] = 'Trabajador'
        ws['C'+rows].style = black_style_border
        ws['C'+rows].alignment = Alignment(horizontal='center')
        ws['D'+rows] = 'Cargo'
        ws['D'+rows].style = black_style_border
        ws['D'+rows].alignment = Alignment(horizontal='center')
        ws['E'+rows] = 'Año'
        ws['E'+rows].style = black_style_border
        ws['E'+rows].alignment = Alignment(horizontal='center')
        ws['F'+rows] = 'Mes'
        ws['F'+rows].style = black_style_border
        ws['F'+rows].alignment = Alignment(horizontal='center')
        ws['G'+rows] = 'Tipo Concepto'
        ws['G'+rows].style = black_style_border
        ws['G'+rows].alignment = Alignment(horizontal='center')
        ws['H'+rows] = 'Concepto'
        ws['H'+rows].style = black_style_border
        ws['H'+rows].alignment = Alignment(horizontal='center')
        ws['I'+rows] = 'Monto'
        ws['I'+rows].style = black_style_border
        ws['I'+rows].alignment = Alignment(horizontal='center')


        for payslip in self.slip_ids:
            for line in payslip.line_ids:
                if line.salary_rule_id.appears_on_payslip and line.category_id.code in('BASIC','ALW','DED'):
                    row += 1
                    rows = str(row)
                    ws['A'+rows] = payslip.employee_id.department_id and payslip.employee_id.department_id.name or ''
                    ws['A'+rows].style = normal_style_border  
                    ws['A'+rows].alignment = Alignment(horizontal='center')
                    ws['B'+rows] = payslip.employee_id.code
                    ws['B'+rows].style = normal_style_border  
                    ws['B'+rows].alignment = Alignment(horizontal='center')
                    ws['C'+rows] = payslip.employee_id.name + ' '+payslip.employee_id.last_name+' '+payslip.employee_id.mother_last_name
                    ws['C'+rows].style = normal_style_border  
                    ws['C'+rows].alignment = Alignment(horizontal='center')
                    ws['D'+rows] = payslip.employee_id.job_id and payslip.employee_id.job_id.name or ''
                    ws['D'+rows].style = normal_style_border  
                    ws['D'+rows].alignment = Alignment(horizontal='center')
                    ws['E'+rows] = int(payslip.date_from.year)
                    ws['E'+rows].style = normal_style_border  
                    ws['E'+rows].alignment = Alignment(horizontal='center')
                    ws['F'+rows] = int(payslip.date_from.month)
                    ws['F'+rows].style = normal_style_border
                    ws['F'+rows].alignment = Alignment(horizontal='center')
                    if line.category_id.code in('BASIC','ALW'):
                        var = 'I'
                    else:
                        var='D'
                    ws['G'+rows] = var
                    ws['G'+rows].style = normal_style_border  
                    ws['G'+rows].alignment = Alignment(horizontal='center')
                    ws['H'+rows] = line.name
                    ws['H'+rows].style = normal_style_border  
                    ws['H'+rows].alignment = Alignment(horizontal='center')
                    ws['I'+rows] = line.total
                    ws['I'+rows].style = normal_style_border  
                    ws['I'+rows].alignment = Alignment(horizontal='center')



        name = 'Resumen_Nomina_'+ str(fields.Datetime.now())
        wb.save(RUTA_BASE+"/../static/src/archivos/"+name+".xlsx")
        return {
            'name': "Resumen de Nómina",
            'type': 'ir.actions.act_url',
            'url': self.env['ir.config_parameter'].get_param('web.base.url')+'/hr_payroll_report_summary/static/src/archivos/'+name+'.xlsx',
            'target': 'new',
        } 

    def send_email(self):
        self.ensure_one()
        for slip in self.slip_ids:
            print(slip.employee_id, slip.employee_id.work_email)
            if slip.employee_id and slip.employee_id.work_email:
                template_mail = self.env.ref('hr_payroll_report_summary.email_template_payslip_employee', False)
                attaches = []
                pdf = self.env.ref('hr_payroll_report_summary.report_rep_payrol').sudo().with_context()._render_qweb_pdf([slip.id])[0]
                pdf_reclamo = base64.b64encode(pdf)
                document_vals = {'name': slip.employee_id.name + '.pdf',
                                 'datas': pdf_reclamo,
                                 'store_fname': slip.employee_id.name + '.pdf',
                                 'res_model': slip._name,
                                 'res_id': slip.id,
                                 'type': 'binary',}
                attach_id = self.env['ir.attachment'].create(document_vals).id
                attaches.append(attach_id)
                template_mail.write({'attachment_ids': [(6, 0, attaches)]})
                template_mail.send_mail(slip.id)

class HRPayrollReportSumaryButtonXML(models.Model):
    _inherit = 'hr.payslip'

    date_today = fields.Datetime.now().strftime("%Y-%m-%d")#llamarlo directo 
    time_today = fields.Datetime.now().strftime("%H:%M:%S")#llamarlo directo 
    document_number_RUC = fields.Char(string='Número de Documento')  #este ya está en company.id 
    order_number = fields.Char(string='Número de Orden') # se llama number en planilla 


    def print_xml(self):
        RUTA_BASE = os.path.dirname(os.path.abspath(__file__))
        # create the file structure
        data = ET.Element('ss:Workbook')
        data.set('xmlns:o','urn:schemas-microsoft-com:office:office')
        data.set('xmlns:x','urn:schemas-microsoft-com:office:excel')
        data.set('xmlns:ss','urn:schemas-microsoft-com:office:spreadsheet')
        data.set('xmlns:html','http://www.w3.org/TR/REC-html40')
        properties = ET.SubElement(data, 'o:DocumentProperties')
        Author = ET.SubElement(properties, 'o:Author')
        Author.text = 'SUNAT'
        LastAuthor = ET.SubElement(properties, 'o:LastAuthor')
        LastAuthor.text = 'SUNAT'
        Created = ET.SubElement(properties, 'o:Created')
        Created.text = '2020-09-21T09:08:38Z'
        LastSaved = ET.SubElement(properties, 'o:LastSaved')
        LastSaved.text = '2020-09-21T09:08:38Z'
        Company = ET.SubElement(properties, 'o:Company')
        Company.text = 'Microsoft'
        Version = ET.SubElement(properties, 'o:Version')
        Version.text = '12.00'

        properties = ET.SubElement(data, 'x:ExcelWorkbook')
        WindowHeight = ET.SubElement(properties, 'x:WindowHeight')
        WindowHeight.text = '12585'
        WindowWidth = ET.SubElement(properties, 'x:WindowWidth')
        WindowWidth.text = '28515'
        WindowTopX = ET.SubElement(properties, 'x:WindowTopX')
        WindowTopX.text = '120'
        WindowTopY = ET.SubElement(properties, 'x:WindowTopY')
        WindowTopY.text = '90'
        ProtectStructure = ET.SubElement(properties, 'x:ProtectStructure')
        ProtectStructure.text = 'False'
        ProtectWindows = ET.SubElement(properties, 'x:ProtectWindows')
        ProtectWindows.text = 'False'

        #ESTILOS ------INICIO--------
        Styles = ET.SubElement(data, 'ss:Styles')
        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s62')
        Borders = ET.SubElement(Style, 'ss:Borders')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Left')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Right')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Top')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Font = ET.SubElement(Style, 'ss:Font')
        Font.set('ss:FontName','Calibri')
        Font.set('ss:Family','Swiss')
        Font.set('ss:Size','12')
        Font.set('ss:Color','#000000')
        Font.set('ss:Bold','0')
        Font.set('ss:Italic','0')
        Alignment = ET.SubElement(Style, 'ss:Alignment')
        Alignment.set('ss:Vertical','Bottom')
        Alignment.set('ss:Horizontal','Center')
        Alignment.set('ss:WrapText','1')
        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s63')
        Borders = ET.SubElement(Style, 'ss:Borders')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Bottom')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Left')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Right')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Font = ET.SubElement(Style, 'ss:Font')
        Font.set('ss:FontName','Calibri')
        Font.set('ss:Family','Swiss')
        Font.set('ss:Size','12')
        Font.set('ss:Color','#000000')
        Font.set('ss:Bold','0')
        Font.set('ss:Italic','0')
        Alignment = ET.SubElement(Style, 'ss:Alignment')
        Alignment.set('ss:Vertical','Bottom')
        Alignment.set('ss:Horizontal','Center')
        Alignment.set('ss:WrapText','1')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s64')
        Borders = ET.SubElement(Style, 'ss:Borders')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Bottom')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Left')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Right')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Top')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Font = ET.SubElement(Style, 'ss:Font')
        Font.set('ss:FontName','Calibri')
        Font.set('ss:Family','Swiss')
        Font.set('ss:Size','12')
        Font.set('ss:Color','#000000')
        Font.set('ss:Bold','0')
        Font.set('ss:Italic','0')
        Alignment = ET.SubElement(Style, 'ss:Alignment')
        Alignment.set('ss:Vertical','Bottom')
        Alignment.set('ss:Horizontal','Center')
        Alignment.set('ss:WrapText','1')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s65')
        Borders = ET.SubElement(Style, 'ss:Borders')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Left')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Right')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Top')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Font = ET.SubElement(Style, 'ss:Font')
        Font.set('ss:FontName','Calibri')
        Font.set('ss:Family','Swiss')
        Font.set('ss:Size','12')
        Font.set('ss:Color','#000000')
        Font.set('ss:Bold','0')
        Font.set('ss:Italic','0')
        Interior = ET.SubElement(Style, 'ss:Interior')
        Interior.set('ss:Color','#DBE5F1')
        Interior.set('ss:Pattern','Solid')
        Alignment = ET.SubElement(Style, 'ss:Alignment')
        Alignment.set('ss:Vertical','Bottom')
        Alignment.set('ss:Horizontal','Center')
        Alignment.set('ss:WrapText','1')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s66')
        Borders = ET.SubElement(Style, 'ss:Borders')

        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Left')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Right')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Bottom')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Font = ET.SubElement(Style, 'ss:Font')
        Font.set('ss:FontName','Calibri')
        Font.set('ss:Family','Swiss')
        Font.set('ss:Size','12')
        Font.set('ss:Color','#000000')
        Font.set('ss:Bold','0')
        Font.set('ss:Italic','0')
        Interior = ET.SubElement(Style, 'ss:Interior')
        Interior.set('ss:Color','#DBE5F1')
        Interior.set('ss:Pattern','Solid')
        Alignment = ET.SubElement(Style, 'ss:Alignment')
        Alignment.set('ss:Vertical','Bottom')
        Alignment.set('ss:Horizontal','Center')
        Alignment.set('ss:WrapText','1')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s67')
        Borders = ET.SubElement(Style, 'ss:Borders')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Left')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Right')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Bottom')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Top')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Font = ET.SubElement(Style, 'ss:Font')
        Font.set('ss:FontName','Calibri')
        Font.set('ss:Family','Swiss')
        Font.set('ss:Size','12')
        Font.set('ss:Color','#000000')
        Font.set('ss:Bold','0')
        Font.set('ss:Italic','0')
        Interior = ET.SubElement(Style, 'ss:Interior')
        Interior.set('ss:Color','#DBE5F1')
        Interior.set('ss:Pattern','Solid')
        Alignment = ET.SubElement(Style, 'ss:Alignment')
        Alignment.set('ss:Vertical','Bottom')
        Alignment.set('ss:Horizontal','Center')
        Alignment.set('ss:WrapText','1')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s68')
        Borders = ET.SubElement(Style, 'ss:Borders')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Left')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Right')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Font = ET.SubElement(Style, 'ss:Font')
        Font.set('ss:FontName','Calibri')
        Font.set('ss:Family','Swiss')
        Font.set('ss:Size','12')
        Font.set('ss:Color','#000000')
        Font.set('ss:Bold','0')
        Font.set('ss:Italic','0')
        Interior = ET.SubElement(Style, 'ss:Interior')
        Interior.set('ss:Color','#DBE5F1')
        Interior.set('ss:Pattern','Solid')
        Alignment = ET.SubElement(Style, 'ss:Alignment')
        Alignment.set('ss:Vertical','Bottom')
        Alignment.set('ss:Horizontal','Center')
        Alignment.set('ss:WrapText','1')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s68')
        Borders = ET.SubElement(Style, 'ss:Borders')

        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Left')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Font = ET.SubElement(Style, 'ss:Font')
        Font.set('ss:FontName','Calibri')
        Font.set('ss:Family','Swiss')
        Font.set('ss:Size','12')
        Font.set('ss:Color','#000000')
        Font.set('ss:Bold','0')
        Font.set('ss:Italic','0')
        Alignment = ET.SubElement(Style, 'ss:Alignment')
        Alignment.set('ss:Vertical','Bottom')
        Alignment.set('ss:Horizontal','Center')
        Alignment.set('ss:WrapText','1')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s70')
        Borders = ET.SubElement(Style, 'ss:Borders')

        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Right')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Font = ET.SubElement(Style, 'ss:Font')
        Font.set('ss:FontName','Calibri')
        Font.set('ss:Family','Swiss')
        Font.set('ss:Size','12')
        Font.set('ss:Color','#000000')
        Font.set('ss:Bold','0')
        Font.set('ss:Italic','0')
        Alignment = ET.SubElement(Style, 'ss:Alignment')
        Alignment.set('ss:Vertical','Bottom')
        Alignment.set('ss:Horizontal','Center')
        Alignment.set('ss:WrapText','1')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s71')
        Borders = ET.SubElement(Style, 'ss:Borders')

        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Left')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Font = ET.SubElement(Style, 'ss:Font')
        Font.set('ss:FontName','Calibri')
        Font.set('ss:Family','Swiss')
        Font.set('ss:Size','12')
        Font.set('ss:Color','#000000')
        Font.set('ss:Bold','0')
        Font.set('ss:Italic','0')
        Interior = ET.SubElement(Style, 'ss:Interior')
        Interior.set('ss:Color','#DBE5F1')
        Interior.set('ss:Pattern','Solid')
        Alignment = ET.SubElement(Style, 'ss:Alignment')
        Alignment.set('ss:Vertical','Bottom')
        Alignment.set('ss:Horizontal','Center')
        Alignment.set('ss:WrapText','1')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s72')
        Borders = ET.SubElement(Style, 'ss:Borders')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Right')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Font = ET.SubElement(Style, 'ss:Font')
        Font.set('ss:FontName','Calibri')
        Font.set('ss:Family','Swiss')
        Font.set('ss:Size','12')
        Font.set('ss:Color','#000000')
        Font.set('ss:Bold','0')
        Font.set('ss:Italic','0')
        Interior = ET.SubElement(Style, 'ss:Interior')
        Interior.set('ss:Color','#DBE5F1')
        Interior.set('ss:Pattern','Solid')
        Alignment = ET.SubElement(Style, 'ss:Alignment')
        Alignment.set('ss:Vertical','Bottom')
        Alignment.set('ss:Horizontal','Center')
        Alignment.set('ss:WrapText','1')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s73')
        Borders = ET.SubElement(Style, 'ss:Borders')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Top')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Alignment = ET.SubElement(Style, 'ss:Alignment')
        Alignment.set('ss:Vertical','Bottom')
        Alignment.set('ss:Horizontal','Center')
        Alignment.set('ss:WrapText','1')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s74')
        Borders = ET.SubElement(Style, 'ss:Borders')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Bottom')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Alignment = ET.SubElement(Style, 'ss:Alignment')
        Alignment.set('ss:Vertical','Bottom')
        Alignment.set('ss:Horizontal','Center')
        Alignment.set('ss:WrapText','1')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s75')
        Borders = ET.SubElement(Style, 'ss:Borders')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Left')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Right')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Top')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Font = ET.SubElement(Style, 'ss:Font')
        Font.set('ss:FontName','Calibri')
        Font.set('ss:Family','Swiss')
        Font.set('ss:Size','12')
        Font.set('ss:Color','#000000')
        Font.set('ss:Bold','0')
        Font.set('ss:Italic','0')
        Interior = ET.SubElement(Style, 'ss:Interior')
        Interior.set('ss:Color','#DBE5F1')
        Interior.set('ss:Pattern','Solid')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s76')
        Borders = ET.SubElement(Style, 'ss:Borders')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Left')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Right')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Bottom')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Font = ET.SubElement(Style, 'ss:Font')
        Font.set('ss:FontName','Calibri')
        Font.set('ss:Family','Swiss')
        Font.set('ss:Size','12')
        Font.set('ss:Color','#000000')
        Font.set('ss:Bold','0')
        Font.set('ss:Italic','0')
        Interior = ET.SubElement(Style, 'ss:Interior')
        Interior.set('ss:Color','#DBE5F1')
        Interior.set('ss:Pattern','Solid')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s77')
        Borders = ET.SubElement(Style, 'ss:Borders')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Left')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Right')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Font = ET.SubElement(Style, 'ss:Font')
        Font.set('ss:FontName','Calibri')
        Font.set('ss:Family','Swiss')
        Font.set('ss:Size','12')
        Font.set('ss:Color','#000000')
        Font.set('ss:Bold','0')
        Font.set('ss:Italic','0')
        Interior = ET.SubElement(Style, 'ss:Interior')
        Interior.set('ss:Color','#DBE5F1')
        Interior.set('ss:Pattern','Solid')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s79')
        Borders = ET.SubElement(Style, 'ss:Borders')
        NumberFormat = ET.SubElement(Style, 'ss:NumberFormat')
        NumberFormat.set('ss:Format','Standard')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s80')
        Borders = ET.SubElement(Style, 'ss:Borders')
        NumberFormat = ET.SubElement(Style, 'ss:NumberFormat')
        NumberFormat.set('ss:Format','Standard')
        Interior = ET.SubElement(Style, 'ss:Interior')
        Interior.set('ss:Color','#DBE5F1')
        Interior.set('ss:Pattern','Solid')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s81')
        Borders = ET.SubElement(Style, 'ss:Borders')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Right')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        NumberFormat = ET.SubElement(Style, 'ss:NumberFormat')
        NumberFormat.set('ss:Format','Standard')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s82')
        Borders = ET.SubElement(Style, 'ss:Borders')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Right')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Alignment = ET.SubElement(Style, 'ss:Alignment')
        Alignment.set('ss:Vertical','Bottom')
        Alignment.set('ss:Horizontal','Right')
        Alignment.set('ss:WrapText','1')
        NumberFormat = ET.SubElement(Style, 'ss:NumberFormat')
        NumberFormat.set('ss:Format','Standard')
        Interior = ET.SubElement(Style, 'ss:Interior')
        Interior.set('ss:Color','#DBE5F1')
        Interior.set('ss:Pattern','Solid')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s83')
        Borders = ET.SubElement(Style, 'ss:Borders')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Left')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Font = ET.SubElement(Style, 'ss:Font')
        Font.set('ss:FontName','Calibri')
        Font.set('ss:Family','Swiss')
        Font.set('ss:Size','12')
        Font.set('ss:Color','#000000')
        Font.set('ss:Bold','0')
        Font.set('ss:Italic','0')
        Interior = ET.SubElement(Style, 'ss:Interior')
        Interior.set('ss:Color','#DBE5F1')
        Interior.set('ss:Pattern','Solid')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s78')
        Borders = ET.SubElement(Style, 'ss:Borders')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Left')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Right')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Bottom')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Top')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Font = ET.SubElement(Style, 'ss:Font')
        Font.set('ss:FontName','Calibri')
        Font.set('ss:Family','Swiss')
        Font.set('ss:Size','12')
        Font.set('ss:Color','#000000')
        Font.set('ss:Bold','0')
        Font.set('ss:Italic','0')
        Interior = ET.SubElement(Style, 'ss:Interior')
        Interior.set('ss:Color','#DBE5F1')
        Interior.set('ss:Pattern','Solid')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s84')
        Borders = ET.SubElement(Style, 'ss:Borders')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Left')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Right')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Top')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Font = ET.SubElement(Style, 'ss:Font')
        Font.set('ss:FontName','Calibri')
        Font.set('ss:Family','Swiss')
        Font.set('ss:Size','12')
        Font.set('ss:Color','#000000')
        Font.set('ss:Bold','0')
        Font.set('ss:Italic','0')
        Interior = ET.SubElement(Style, 'ss:Interior')
        Interior.set('ss:Color','#DBE5F1')
        Interior.set('ss:Pattern','Solid')
        Alignment = ET.SubElement(Style, 'ss:Alignment')
        Alignment.set('ss:Vertical','Bottom')
        Alignment.set('ss:Horizontal','Center')
        Alignment.set('ss:WrapText','1')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s85')
        Borders = ET.SubElement(Style, 'ss:Borders')
        Font = ET.SubElement(Style, 'ss:Font')
        Font.set('ss:FontName','Calibri')
        Font.set('ss:Family','Swiss')
        Font.set('ss:Size','12')
        Font.set('ss:Color','#000000')
        Font.set('ss:Bold','1')
        Font.set('ss:Italic','0')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s86')
        Borders = ET.SubElement(Style, 'ss:Borders')
        Font = ET.SubElement(Style, 'ss:Font')
        Font.set('ss:FontName','Calibri')
        Font.set('ss:Family','Swiss')
        Font.set('ss:Size','8')
        Font.set('ss:Color','#000000')
        Font.set('ss:Bold','0')
        Font.set('ss:Italic','1')

        Style = ET.SubElement(Styles, 'ss:Style')
        Style.set('ss:ID','s87')
        Borders = ET.SubElement(Style, 'ss:Borders')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Left')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Right')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Top')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Border = ET.SubElement(Borders, 'ss:Border')
        Border.set('ss:Position','Bottom')
        Border.set('ss:LineStyle','Continuous')
        Border.set('ss:Weight','1')
        Font = ET.SubElement(Style, 'ss:Font')
        Font.set('ss:FontName','Calibri')
        Font.set('ss:Family','Swiss')
        Font.set('ss:Size','12')
        Font.set('ss:Color','#000000')
        Font.set('ss:Bold','0')
        Font.set('ss:Italic','0')
        Interior = ET.SubElement(Style, 'ss:Interior')
        Interior.set('ss:Color','#DBE5F1')
        Interior.set('ss:Pattern','Solid')
        Alignment = ET.SubElement(Style, 'ss:Alignment')
        Alignment.set('ss:Vertical','Bottom')
        Alignment.set('ss:Horizontal','Center')
        Alignment.set('ss:WrapText','1')
        #ESTILOS ------FIN--------

        #HOJA DE TRABAJO ------INICIO--------
        Worksheet = ET.SubElement(data, 'ss:Worksheet')


        Table = ET.SubElement(Worksheet, 'ss:Table')
        Table.set('ss:ExpandedColumnCount','10')
        Table.set('ss:ExpandedRowCount','78')
        Table.set('x:FullColumns','1')
        Table.set('x:FullRows','1')
        Table.set('ss:DefaultColumnWidth','60')
        Table.set('ss:DefaultRowHeight','15')

        #ROW ------INICIO--------
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index','2')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s85')
        Cell.set('ss:MergeAcross','3')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'R08: Trabajador – Datos de boleta de pago'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','9')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = str(self.date_today)

        #ROW ------FIN--------

        #ROW ------INICIO--------
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index','3')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:MergeAcross','3')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = '(Contiene datos mínimos de una Boleta de Pago)'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','9')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = str(self.time_today)
        #ROW ------FIN--------

        #ROW ------INICIO--------
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index','6')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s75')
        Cell.set('ss:MergeAcross','7')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'RUC : '+str(self.company_id.partner_id.vat) or ''

        #ROW ------FIN--------

        #ROW ------INICIO--------
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index','7')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s77')
        Cell.set('ss:MergeAcross','7')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Empleador : '+str(self.company_id.name) or ''
        #ROW ------FIN--------

        #ROW ------INICIO--------
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index','8')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s77')
        Cell.set('ss:MergeAcross','7')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Periodo : '+str(self.date_from)

        #ROW ------FIN--------

        #ROW ------INICIO--------
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index','9')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s76')
        Cell.set('ss:MergeAcross','7')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'PDT Planilla Electrónica - PLAME            Número de Orden : '+str(self.order_number)

        #ROW ------FIN--------

        #ROW ------INICIO--------
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index','11')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s84')
        Cell.set('ss:MergeAcross','1')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Documento de Identidad'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','4')
        Cell.set('ss:StyleID','s65')
        Cell.set('ss:MergeAcross','3')
        Cell.set('ss:MergeDown','1')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Nombre y Apellidos'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','8')
        Cell.set('ss:StyleID','s65')
        Cell.set('ss:MergeAcross','1')
        Cell.set('ss:MergeDown','1')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Situación'
        #ROW ------FIN--------

        #ROW ------INICIO--------
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index','12')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s65')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Tipo'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','3')
        Cell.set('ss:StyleID','s65')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Número'
        #ROW ------FIN--------

        #ROW ------INICIO--------
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index','13')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s62')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = str(self.employee_id.type_document_id.name  or '')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','3')
        Cell.set('ss:StyleID','s62')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = str(self.employee_id.identification_id  or ' ')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','4')
        Cell.set('ss:StyleID','s62')
        Cell.set('ss:MergeAcross','3')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = str(self.employee_id.name)+' '+str(self.employee_id.last_name or '')+' '+str(self.employee_id.mother_last_name or '')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','8')
        Cell.set('ss:StyleID','s62')
        Cell.set('ss:MergeAcross','1')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = str(self.employee_id.special_situation.name  or ' ')
        #ROW ------FIN--------

        #ROW ------INICIO--------
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index','14')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s65')
        Cell.set('ss:MergeAcross','1')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Fecha de Ingreso'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','4')
        Cell.set('ss:StyleID','s65')
        Cell.set('ss:MergeAcross','1')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Tipo de Trabajador'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','6')
        Cell.set('ss:StyleID','s65')
        Cell.set('ss:MergeAcross','1')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Regimen Pensionario'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','8')
        Cell.set('ss:StyleID','s65')
        Cell.set('ss:MergeAcross','1')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'CUSPP'

        #ROW ------FIN--------

        #ROW ------INICIO--------
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index','15')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s62')
        Cell.set('ss:MergeAcross','1')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = str(self.contract_id.date_start or ' ')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','4')
        Cell.set('ss:StyleID','s62')
        Cell.set('ss:MergeAcross','1')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = str(self.employee_id.type_employee.name or ' ') 

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','6')
        Cell.set('ss:StyleID','s62')
        Cell.set('ss:MergeAcross','1')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = str(self.employee_id.pension_fund.name or ' ')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','8')
        Cell.set('ss:StyleID','s62')
        Cell.set('ss:MergeAcross','1')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = str(self.employee_id.code_cussp or ' ')
        #ROW ------FIN--------

        #ROW ------INICIO--------

        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index','16')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s65')
        Cell.set('ss:MergeDown','1')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Días Laborados'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','3')
        Cell.set('ss:StyleID','s65')
        Cell.set('ss:MergeDown','1')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Días No Laborados'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','4')
        Cell.set('ss:StyleID','s65')
        Cell.set('ss:MergeDown','1')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Días subsidiados'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','5')
        Cell.set('ss:StyleID','s65')
        Cell.set('ss:MergeDown','1')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Condición'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','6')
        Cell.set('ss:StyleID','s65')
        Cell.set('ss:MergeAcross','1')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Jornada Ordinaria'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','8')
        Cell.set('ss:StyleID','s65')
        Cell.set('ss:MergeAcross','1')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Sobretiempo'

        #ROW ------FIN--------

        #ROW ------INICIO--------
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index','17')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','6')
        Cell.set('ss:StyleID','s65')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Total Horas'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','7')
        Cell.set('ss:StyleID','s65')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Minutos'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','8')
        Cell.set('ss:StyleID','s65')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Total Horas'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','9')
        Cell.set('ss:StyleID','s65')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Minutos'

        #ROW ------FIN--------
        total_dias_trab = 0
        total_horas_trab = 0
        total_no_lab = 0
        total_horas_no_lab = 0
        total_horas_extras = 0
        total_subs = 0


        dt = self.env.ref('hr_tareo.work_entry_type_diat')
        falt = self.env.ref('hr_tareo.work_entry_type_falt')
        lsg = self.env.ref('hr_tareo.work_entry_type_csgo')

        he25 = self.env.ref('hr_tareo.work_entry_type_he25')
        he35 = self.env.ref('hr_tareo.work_entry_type_he35')
        
        dmed = self.env.ref('hr_tareo.work_entry_type_dmed')
        subm = self.env.ref('hr_tareo.work_entry_type_subm')
        sube = self.env.ref('hr_tareo.work_entry_type_sube')
        lcgo = self.env.ref('hr_tareo.work_entry_type_lcgo')

        for entry in self.worked_days_line_ids:
            print(entry.work_entry_type_id)
            if entry.work_entry_type_id.id == dt.id:
                total_dias_trab += entry.number_of_days
                total_horas_trab += entry.number_of_hours
            if entry.work_entry_type_id.id in [falt.id, lsg.id]:
                total_no_lab += entry.number_of_days
                total_horas_no_lab += entry.number_of_hours
            if entry.work_entry_type_id.id in [he25.id, he35.id]:
                total_horas_extras+= entry.number_of_hours

            if entry.work_entry_type_id.id in [dmed.id,subm.id,sube.id,lcgo.id,]:
                total_subs += entry.number_of_days
        #ROW ------INICIO--------
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index','18')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s62')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = str(total_dias_trab)

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','3')
        Cell.set('ss:StyleID','s62')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = str(total_no_lab)

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','4')
        Cell.set('ss:StyleID','s62')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = str(total_subs)

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','5')
        Cell.set('ss:StyleID','s62')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Especificar' #condicion

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','6')
        Cell.set('ss:StyleID','s62')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = str(total_horas_trab)

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','7')
        Cell.set('ss:StyleID','s62')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 0 #minutos

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','8')
        Cell.set('ss:StyleID','s62')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = str(total_horas_extras)

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','9')
        Cell.set('ss:StyleID','s62')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 0 #minutos

        #ROW ------FIN--------

        #ROW ------INICIO--------
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index','19')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s65')
        Cell.set('ss:MergeAcross','5')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Motivo de Suspensión de Labores'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','8')
        Cell.set('ss:StyleID','s65')
        Cell.set('ss:MergeAcross','1')
        Cell.set('ss:MergeDown','1')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Otros empleadores por Rentas de 5ta categoría'

        #ROW ------FIN--------

        #ROW ------INICIO--------
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index','20')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s65')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Tipo'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','3')
        Cell.set('ss:StyleID','s65')
        Cell.set('ss:MergeAcross','3')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Motivo'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','7')
        Cell.set('ss:StyleID','s65')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'N.º Días'

        #ROW ------FIN--------

        #ROW ------INICIO--------
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index','21')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s64')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = '-'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','3')
        Cell.set('ss:StyleID','s64')
        Cell.set('ss:MergeAcross','3')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = '-'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','7')
        Cell.set('ss:StyleID','s64')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = '-'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','8')
        Cell.set('ss:StyleID','s64')
        Cell.set('ss:MergeAcross','1')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = '-'

        #ROW ------FIN--------

        #ROW ------INICIO--------
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index','23')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s87')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Código'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','3')
        Cell.set('ss:StyleID','s87')
        Cell.set('ss:MergeAcross','3')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Conceptos'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','7')
        Cell.set('ss:StyleID','s87')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Ingresos S/.'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','8')
        Cell.set('ss:StyleID','s87')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Descuentos S/.'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','9')
        Cell.set('ss:StyleID','s87')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Neto S/.'

        #ROW ------FIN--------

        #ROW ------INICIO--------
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index','24')

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s77')
        Cell.set('ss:MergeAcross','7')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Ingresos'
        #ROW ------FIN--------
        index = 25
        amount_total = 0
        cat_basic_id = self.env.ref('hr_payroll.BASIC')
        cat_alw_id = self.env.ref('hr_payroll.ALW')

        for concepto in self.env['hr.payslip.line'].search([('slip_id', '=', self.id),
            ('salary_rule_id.category_id', 'in', [cat_basic_id.id, cat_alw_id.id])]):
            Row = ET.SubElement(Table, 'ss:Row')
            Row.set('ss:Index',str(index))

            Cell = ET.SubElement(Row, 'ss:Cell')
            Cell.set('ss:Index','2')
            Cell.set('ss:StyleID','s69')

            Data = ET.SubElement(Cell, 'ss:Data')
            Data.set('ss:Type','String')
            Data.text = str(concepto.code)

            Cell = ET.SubElement(Row, 'ss:Cell')
            Cell.set('ss:Index','3')
            Cell.set('ss:MergeAcross','3')

            Data = ET.SubElement(Cell, 'ss:Data')
            Data.set('ss:Type','String')
            Data.text = str(concepto.name)

            Cell = ET.SubElement(Row, 'ss:Cell')
            Cell.set('ss:Index','7')
            Cell.set('ss:StyleID','s79')

            Data = ET.SubElement(Cell, 'ss:Data')
            Data.set('ss:Type','Number')
            Data.text = str(concepto.amount)

            Cell = ET.SubElement(Row, 'ss:Cell')
            Cell.set('ss:Index','8')
            Cell.set('ss:StyleID','s70')
            Cell.set('ss:MergeAcross','1')
            amount_total += concepto.amount
            index +=1
        #ROW ------FIN--------

        #ROW ------INICIO--------
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index',str(index))

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s77')
        Cell.set('ss:MergeAcross','7')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Descuentos'
        #ROW ------FIN--------
        index +=1
        cat_ded_id = self.env.ref('hr_payroll.DED')

        for concepto in self.env['hr.payslip.line'].search([('slip_id', '=', self.id),
                     ('salary_rule_id.category_id', 'in', [cat_ded_id.id])]):
            
            Row = ET.SubElement(Table, 'ss:Row')
            Row.set('ss:Index',str(index))

            Cell = ET.SubElement(Row, 'ss:Cell')
            Cell.set('ss:Index','2')
            Cell.set('ss:StyleID','s69')

            Data = ET.SubElement(Cell, 'ss:Data')
            Data.set('ss:Type','String')
            Data.text = str(concepto.code)

            Cell = ET.SubElement(Row, 'ss:Cell')
            Cell.set('ss:Index','3')
            Cell.set('ss:MergeAcross','3')

            Data = ET.SubElement(Cell, 'ss:Data')
            Data.set('ss:Type','String')
            Data.text = str(concepto.name)

            Cell = ET.SubElement(Row, 'ss:Cell')
            Cell.set('ss:Index','8')
            Cell.set('ss:StyleID','s79')

            Data = ET.SubElement(Cell, 'ss:Data')
            Data.set('ss:Type','Number')
            Data.text = str(concepto.amount)

            Cell = ET.SubElement(Row, 'ss:Cell')
            Cell.set('ss:Index','9')
            Cell.set('ss:StyleID','s70')
            amount_total -= concepto.amount
            index +=1
        #ROW ------FIN--------
        #ROW ------INICIO--------
        """
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index',str(index))

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s77')
        Cell.set('ss:MergeAcross','7')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Aportes del Trabajador'
        #ROW ------FIN--------
        index +=1
        for concepto in self.env['hr.payslip.line'].search([('slip_id', '=', self.id),
                                 ('salary_rule_id.type_concept', '=', 'worker_contribution')]):
            Row = ET.SubElement(Table, 'ss:Row')
            Row.set('ss:Index',str(index))

            Cell = ET.SubElement(Row, 'ss:Cell')
            Cell.set('ss:Index','2')
            Cell.set('ss:StyleID','s69')

            Data = ET.SubElement(Cell, 'ss:Data')
            Data.set('ss:Type','String')
            Data.text = str(concepto.code)

            Cell = ET.SubElement(Row, 'ss:Cell')
            Cell.set('ss:Index','3')
            Cell.set('ss:MergeAcross','3')

            Data = ET.SubElement(Cell, 'ss:Data')
            Data.set('ss:Type','String')
            Data.text = str(concepto.name)

            Cell = ET.SubElement(Row, 'ss:Cell')
            Cell.set('ss:Index','8')
            Cell.set('ss:StyleID','s79')

            Data = ET.SubElement(Cell, 'ss:Data')
            Data.set('ss:Type','Number')
            Data.text = str(concepto.amount)

            Cell = ET.SubElement(Row, 'ss:Cell')
            Cell.set('ss:Index','9')
            Cell.set('ss:StyleID','s70')
            amount_total -= concepto.amount
            index +=1
        """
        #ROW ------FIN--------

        #ROW ------INICIO--------
        
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index',str(index))

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s83')
        Cell.set('ss:MergeAcross','6')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Neto a Pagar'

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','9')
        Cell.set('ss:StyleID','s82')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','Number')
        #Data.text = str(amount_total)
        cat_net_id = self.env.ref('hr_payroll.NET')

        Data.text = str(self.env['hr.payslip.line'].search([('slip_id', '=', self.id),
                     ('salary_rule_id.category_id', 'in', [cat_net_id.id])], limit=1).amount)
        #ROW ------FIN--------

        #ROW ------INICIO--------
        index +=1
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index',str(index))

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s73')
        Cell.set('ss:MergeAcross','7')

        #ROW ------FIN--------

        #ROW ------INICIO--------
        index +=1
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index',str(index))

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s78')
        Cell.set('ss:MergeAcross','7')

        Data = ET.SubElement(Cell, 'ss:Data')
        Data.set('ss:Type','String')
        Data.text = 'Aportes de Empleador'

        #ROW ------FIN--------

        #ROW ------INICIO--------
        index +=1
        cat_comp_id = self.env.ref('hr_payroll.COMP')

        for concepto in self.env['hr.payslip.line'].search([('slip_id', '=', self.id),
                                 ('salary_rule_id.category_id', 'in', [cat_comp_id.id])]):
            Row = ET.SubElement(Table, 'ss:Row')
            Row.set('ss:Index',str(index))

            Cell = ET.SubElement(Row, 'ss:Cell')
            Cell.set('ss:Index','2')
            Cell.set('ss:StyleID','s69')

            Data = ET.SubElement(Cell, 'ss:Data')
            Data.set('ss:Type','String')
            Data.text = str(concepto.code)

            Cell = ET.SubElement(Row, 'ss:Cell')
            Cell.set('ss:Index','3')
            Cell.set('ss:MergeAcross','3')

            Data = ET.SubElement(Cell, 'ss:Data')
            Data.set('ss:Type','String')
            Data.text = str(concepto.name)

            Cell = ET.SubElement(Row, 'ss:Cell')
            Cell.set('ss:Index','9')
            Cell.set('ss:StyleID','s81')

            Data = ET.SubElement(Cell, 'ss:Data')
            Data.set('ss:Type','Number')
            Data.text = str(concepto.amount)
            index +=1

        #ROW ------FIN--------

        #ROW ------INICIO--------
        
        Row = ET.SubElement(Table, 'ss:Row')
        Row.set('ss:Index',str(index))

        Cell = ET.SubElement(Row, 'ss:Cell')
        Cell.set('ss:Index','2')
        Cell.set('ss:StyleID','s73')
        Cell.set('ss:MergeAcross','7')
        #ROW ------FIN--------


        WorksheetOptions = ET.SubElement(Worksheet, 'ss:WorksheetOptions')
        WorksheetOptions.set('xmlns','urn:schemas-microsoft-com:office:excel')

        PageSetup = ET.SubElement(WorksheetOptions, 'PageSetup')

        Header = ET.SubElement(PageSetup, 'Header')
        Header.set('x:Margin','0.2')
        Header.set('x:Data','&amp;Z&amp;&quot;Calibri&quot;&amp;10Superintendencia Nacional de Aduanas y Administración Tributaria&amp;DPágina: &amp;#')
        Footer = ET.SubElement(PageSetup, 'Footer')
        Footer.set('x:Margin','0.2')
        Footer.set('x:Data','Generado por el PDT Planilla Electrónica PLAME')
        PageMargins = ET.SubElement(PageSetup, 'PageMargins')
        PageMargins.set('x:Top','0.74')
        PageMargins.set('x:Right','0.74')
        PageMargins.set('x:Bottom','0.98')
        PageMargins.set('x:Left','0.98')

        Print = ET.SubElement(WorksheetOptions, 'Print')

        ValidPrinterInfo = ET.SubElement(Print, 'ValidPrinterInfo')
        PaperSizeIndex = ET.SubElement(Print, 'PaperSizeIndex')
        PaperSizeIndex.text = '9'
        Scale = ET.SubElement(Print, 'Scale')
        Scale.text = '80'
        HorizontalResolution = ET.SubElement(Print, 'HorizontalResolution')
        HorizontalResolution.text = '600'
        VerticalResolution = ET.SubElement(Print, 'VerticalResolution')
        VerticalResolution.text = '600'

        name = self.employee_id.name+'.xml'

        mydata = ET.tostring(data)
        myfile = open(RUTA_BASE+"/../static/src/archivos/"+name, "wb")
        myfile.write(mydata)



        return {
        'name': "Resumen de Nómina",
        'type': 'ir.actions.act_url',
        'url': self.env['ir.config_parameter'].get_param('web.base.url')+'/hr_payroll_report_summary/static/src/archivos/'+name,
        'target': 'new',
        }

