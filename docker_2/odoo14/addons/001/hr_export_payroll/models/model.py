
# -- coding: utf-8 --
from odoo import models, fields, api
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Protection, Font
import os
from odoo.exceptions import ValidationError
import xml.etree.ElementTree as ET
import datetime

class HRContractInherit(models.Model):
    _inherit = "hr.payslip.run" 

    def export_report_payroll(self):
        RUTA_BASE = os.path.dirname(os.path.abspath(__file__))
        wb = load_workbook(RUTA_BASE+'/plantilla_reporte.xlsx')
        sheet_ranges = wb['Hoja1']
        ws = wb['Hoja1']
        ws.title = self.company_id.name
        bd = Side(style='thin', color="000000")
        normal_style = NamedStyle(name="normal_style")
        normal_style.font = Font(name='Arial Narrow', bold=False, size=10)
        normal_style.alignment = Alignment(horizontal='right')
        wb.add_named_style(normal_style)
        normal_style_border = NamedStyle(name="normal_style_border")
        normal_style_border.font = Font(name='Arial Narrow', bold=False, size=10)
        normal_style_border.alignment = Alignment(horizontal='right')
        normal_style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        wb.add_named_style(normal_style_border)
        black_style = NamedStyle(name="black_style")
        black_style.font = Font(name='Arial Narrow', bold=True, size=10)
        black_style.alignment = Alignment(horizontal='right')
        wb.add_named_style(black_style)
        black_style_border = NamedStyle(name="black_style_border")
        black_style_border.font = Font(name='Arial Narrow', bold=True, size=10)
        black_style_border.alignment = Alignment(horizontal='right')
        black_style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        wb.add_named_style(black_style_border)
        row=8
        rows = str(row)
        ws['A'+rows] = 'CEDULA°'
        ws['A'+rows].style = black_style_border
        ws['A'+rows].alignment = Alignment(horizontal='center')
        ws['B'+rows] = 'NOMBRE'
        ws['B'+rows].style = black_style_border
        ws['B'+rows].alignment = Alignment(horizontal='center')
        ws['C'+rows] = 'SALARIO BÁSICO'
        ws['C'+rows].style = black_style_border
        ws['C'+rows].alignment = Alignment(horizontal='center')
        ws['D'+rows] = 'VACACIONES DISFRUTADAS'
        ws['D'+rows].style = black_style_border
        ws['D'+rows].alignment = Alignment(horizontal='center')
        ws['E'+rows] = 'CUOTA SOSTENIMIENTO'
        ws['E'+rows].style = black_style_border
        ws['E'+rows].alignment = Alignment(horizontal='center')
        ws['F'+rows] = 'COMISIONES'
        ws['F'+rows].style = black_style_border
        ws['F'+rows].alignment = Alignment(horizontal='center')
        ws['G'+rows] = 'HORAS EXTRAS DIURNAS'
        ws['G'+rows].style = black_style_border
        ws['G'+rows].alignment = Alignment(horizontal='center')
        ws['H'+rows] = 'HORAS EXTRAS NOCTURNAS'
        ws['H'+rows].style = black_style_border
        ws['H'+rows].alignment = Alignment(horizontal='center')
        ws['I'+rows] = 'RECARGO NOCTURNO'
        ws['I'+rows].style = black_style_border
        ws['I'+rows].alignment = Alignment(horizontal='center')
        ws['J'+rows] = 'RECARGO NOCTURNO FESTIVO'
        ws['J'+rows].style = black_style_border
        ws['J'+rows].alignment = Alignment(horizontal='center')
        ws['K'+rows] = 'HORAS FESTIVAS NOCTURNAS'
        ws['K'+rows].style = black_style_border
        ws['K'+rows].alignment = Alignment(horizontal='center')
        ws['L'+rows] = 'DOMINICALES CON COMPENSATORIO'
        ws['L'+rows].style = black_style_border
        ws['L'+rows].alignment = Alignment(horizontal='center')
        ws['M'+rows] = 'AUXILIO DE ALIMENTACIÓN'
        ws['M'+rows].style = black_style_border
        ws['M'+rows].alignment = Alignment(horizontal='center')
        ws['N'+rows] = 'AUXILIO MOVILIZACIÓN'
        ws['N'+rows].style = black_style_border
        ws['N'+rows].alignment = Alignment(horizontal='center')
        ws['O'+rows] = 'AUXILIO ALI-VIVIENDA'
        ws['O'+rows].style = black_style_border
        ws['O'+rows].alignment = Alignment(horizontal='center')
        ws['P'+rows] = 'SUBSIDIO DE TRANSPORTE'
        ws['P'+rows].style = black_style_border
        ws['P'+rows].alignment = Alignment(horizontal='center')
        ws['Q'+rows] = 'SALUD'
        ws['Q'+rows].style = black_style_border
        ws['Q'+rows].alignment = Alignment(horizontal='center')
        ws['R'+rows] = 'PENSIÓN'
        ws['R'+rows].style = black_style_border
        ws['R'+rows].alignment = Alignment(horizontal='center')
        ws['S'+rows] = 'FONDO DE SOLIDARIDAD'
        ws['S'+rows].style = black_style_border
        ws['S'+rows].alignment = Alignment(horizontal='center')
        ws['T'+rows] = 'FONDO DE SUBSISTENCIA'
        ws['T'+rows].style = black_style_border
        ws['T'+rows].alignment = Alignment(horizontal='center')
        ws['U'+rows] = 'RETENCIÓN EN LA FUENTE'
        ws['U'+rows].style = black_style_border
        ws['U'+rows].alignment = Alignment(horizontal='center')
        ws['V'+rows] = 'LIBRANZA'
        ws['V'+rows].style = black_style_border
        ws['V'+rows].alignment = Alignment(horizontal='center')
        ws['W'+rows] = 'APORTES FONDO DE PENSIONES VOLUNTARIAS'
        ws['W'+rows].style = black_style_border
        ws['W'+rows].alignment = Alignment(horizontal='center')
        ws['X'+rows] = 'DESCUENTOS A EMPLEADOS'
        ws['X'+rows].style = black_style_border
        ws['X'+rows].alignment = Alignment(horizontal='center')
        ws['Y'+rows] = 'APORTES CUENTAS AFC'
        ws['Y'+rows].style = black_style_border
        ws['Y'+rows].alignment = Alignment(horizontal='center')
        ws['Z'+rows] = 'Neto a Pagar'
        ws['Z'+rows].style = black_style_border
        ws['Z'+rows].alignment = Alignment(horizontal='center')
        
        sequence=0
        for payslip in self.slip_ids:
            sequence += 1
            row=row+1
            rows = str(row)
            ws['A'+rows] = sequence
            ws['A'+rows].style = normal_style_border
            ws['A'+rows].alignment = Alignment(horizontal='right')
            ws['B'+rows] = payslip.employee_id.name
            ws['B'+rows].style = black_style_border
            ws['B'+rows].alignment = Alignment(horizontal='center')
            ws['C'+rows] = payslip.contract_id.wage
            ws['C'+rows].style = black_style_border
            ws['C'+rows].alignment = Alignment(horizontal='center')
            ws['D'+rows] = (payslip.employee_id.GetBaseVacDisfrutadas(payslip.date_from,payslip.contract_id)+payslip.contract_id.wage)/30
            ws['D'+rows].style = normal_style_border
            ws['D'+rows].alignment = Alignment(horizontal='center')
            ws['E'+rows] = ''
            ws['E'+rows].style = normal_style_border
            ws['E'+rows].alignment = Alignment(horizontal='center')
            ws['F'+rows] = payslip.contract_id.wage
            ws['F'+rows].style = normal_style_border
            ws['F'+rows].alignment = Alignment(horizontal='center')
            ws['G'+rows] = ''
            ws['G'+rows].style = normal_style_border
            ws['G'+rows].alignment = Alignment(horizontal='center')
            ws['H'+rows] = ''
            ws['H'+rows].style = normal_style_border
            ws['H'+rows].alignment = Alignment(horizontal='center')
            ws['I'+rows] = ''
            ws['I'+rows].style = normal_style_border
            ws['I'+rows].alignment = Alignment(horizontal='center')
            ws['J'+rows] = 'SI'
            ws['J'+rows].style = normal_style_border
            ws['J'+rows].alignment = Alignment(horizontal='center')
            ws['K'+rows] = ''
            ws['K'+rows].style = normal_style_border
            ws['K'+rows].alignment = Alignment(horizontal='center')
            ws['L'+rows] = ''
            ws['L'+rows].style = normal_style_border
            ws['L'+rows].alignment = Alignment(horizontal='center')
            ws['M'+rows] = ''
            ws['M'+rows].style = normal_style_border
            ws['M'+rows].alignment = Alignment(horizontal='center')
            ws['N'+rows] = ''
            ws['N'+rows].style = normal_style_border
            ws['N'+rows].alignment = Alignment(horizontal='center')
            ws['O'+rows] = ''
            ws['O'+rows].style = normal_style_border
            ws['O'+rows].alignment = Alignment(horizontal='center')
            ws['P'+rows] = ''
            ws['P'+rows].style = normal_style_border
            ws['P'+rows].alignment = Alignment(horizontal='center')
            ws['Q'+rows] = ''
            ws['Q'+rows].style = normal_style_border
            ws['Q'+rows].alignment = Alignment(horizontal='center')
            ws['R'+rows] = ''
            ws['R'+rows].style = normal_style_border
            ws['R'+rows].alignment = Alignment(horizontal='center')
            ws['S'+rows] = 'FONDO DE SOLIDARIDAD'
            ws['S'+rows].style = black_style_border
            ws['S'+rows].alignment = Alignment(horizontal='center')
            ws['T'+rows] = 'FONDO DE SUBSISTENCIA'
            ws['T'+rows].style = black_style_border
            ws['T'+rows].alignment = Alignment(horizontal='center')
            ws['U'+rows] = 'RETENCION EN LA FUENTE'
            ws['U'+rows].style = black_style_border
            ws['U'+rows].alignment = Alignment(horizontal='center')
            ws['V'+rows] = 'LIBRANZA'
            ws['V'+rows].style = black_style_border
            ws['V'+rows].alignment = Alignment(horizontal='center')
            ws['W'+rows] = 'APORTES FONDO DE PENSIONES VOLUNTARIAS'
            ws['W'+rows].style = black_style_border
            ws['W'+rows].alignment = Alignment(horizontal='center')
            ws['X'+rows] = 'DESCUENTOS A EMPLEADOS'
            ws['X'+rows].style = black_style_border
            ws['X'+rows].alignment = Alignment(horizontal='center')
            ws['Y'+rows] = 'APORTES CUENTAS AFC'
            ws['Y'+rows].style = black_style_border
            ws['Y'+rows].alignment = Alignment(horizontal='center')
            ws['Z'+rows] = 'Neto a Pagar'
            ws['Z'+rows].style = black_style_border
            ws['Z'+rows].alignment = Alignment(horizontal='center')
            
        # Total a pagar
        source = wb[self.company_id.name]
        ws = wb.copy_worksheet(source)
        ws.title = 'Total a Pagar'
        row=1
        rows = str(row)
        ws['A'+rows] = 'EMPRESA: ' +self.company_id.partner_id.name or ''
        ws['A'+rows].style = black_style
        ws['A'+rows].alignment = Alignment(horizontal='left')
        row=2
        rows = str(row)
        ws['A'+rows] = 'RUC:        ' 
        ws['A'+rows].style = black_style
        ws['A'+rows].alignment = Alignment(horizontal='left')
        row=5
        rows = str(row)
        ws['B'+rows] = self.name
        ws['B'+rows].style = black_style
        ws['B'+rows].alignment = Alignment(horizontal='center')
        row=6
        rows = str(row)
        ws['A'+rows] = 'N°'
        ws['A'+rows].style = black_style_border
        ws['A'+rows].alignment = Alignment(horizontal='center')
        ws['B'+rows] = 'COLABORADOR'
        ws['B'+rows].style = black_style_border
        ws['B'+rows].alignment = Alignment(horizontal='center')
        ws['C'+rows] = 'DNI'
        ws['C'+rows].style = black_style_border
        ws['C'+rows].alignment = Alignment(horizontal='center')
        ws['D'+rows] = 'NETO A PAGAR'
        ws['D'+rows].style = black_style_border
        ws['D'+rows].alignment = Alignment(horizontal='center')
        sequence=0
        for payslip in self.slip_ids:
            sequence += 1
            row=row+1
            rows = str(row)
            ws['A'+rows] = sequence
            ws['A'+rows].style = black_style_border
            ws['A'+rows].alignment = Alignment(horizontal='center')
            ws['B'+rows] = payslip.employee_id.name
            ws['B'+rows].style = black_style_border
            ws['B'+rows].alignment = Alignment(horizontal='left')
            ws['C'+rows] = payslip.employee_id.identification_id
            ws['C'+rows].style = normal_style_border
            ws['C'+rows].alignment = Alignment(horizontal='center')
            ws['D'+rows] = payslip.line_ids[-1].total
            ws['D'+rows].style = normal_style_border
        row=row+1
        rows = str(row)
        ws['D'+rows] = '=SUM(D7:D'+str(row-1)+')'
        ws['D'+rows].style = black_style
        name = 'Resumen_Nomina_'+ str(fields.Datetime.now())
        wb.save(RUTA_BASE+"/../static/src/archivos/"+name+".xlsx")
        return {
            'name': "Descargar Reporte",
            'type': 'ir.actions.act_url',
            'url': self.env['ir.config_parameter'].get_param('web.base.url')+'/hr_export_payroll/static/src/archivos/'+name+'.xlsx',
            'target': 'new',
        }
    



    
